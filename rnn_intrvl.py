import numpy as np
import tensorflow as tf
import pandas as pd
import os
import pathlib
# import s_arima_mod as sam
import tensorflow.contrib.eager as tfe
from tensorflow.contrib import rnn
from tensorflow.contrib import seq2seq
from tensorflow.contrib import cudnn_rnn
from tensorflow.nn import rnn_cell

# from tensorflow.python import debug as tf_debug

# import make_slide_win as msw
#import cufflinks as cf
import matplotlib.pyplot as plt
#import matplotlib as mpl
import scipy.stats as spy
import openpyxl
# import numba
# import ppp
# import plotting
# from numba import cuda
from scipy.stats import norm
#import pdb; pdb.set_trace()
from sklearn.utils import shuffle
from sklearn.model_selection import train_test_split
#start = time.time()

learning_rate = 0.01
# when attention,learning_rate must be 0.001
learning_data_day_len = 50
input_digits = 20
output_digits = 20
n_hidden = 40
epochs = 1000
eary_stop_epoch = 150
batch_size = 30
attention_layer_size = 10
num_units = 50
ample = 0
# day = 'Tue'
# learning_length = 700
thrd = 54.5
tchr_frcng_thr = 0.8

tf.reset_default_graph()

# tfe.enable_eager_execution()

input_data_path='/tmp/rnn_input/log_time.csv'
path_output_data='/tmp/rnn_output/'
# path_fig = '/tmp/RNN_python/figures_seq2seq_test/'
# path_output_data = '/tmp/RNN_python/schd_sam_output_data/'
# LOG_DIR = '/tmp/RNN_python/rnn_log'

os.makedirs(path_output_data, exist_ok=True)

class TimeSeriesDataSet:
#時系列データの時間の設定
    def __init__(self, dataframe):
        self.feature_count = len(dataframe.columns)
        self.series_length = len(dataframe)
        self.series_data = dataframe.astype('float32')

    def __getitem__(self, n):
        return TimeSeriesDataSet(self.series_data[n])

    def __length_of_sequences__(self):
        return len(self.series_data)

    def times(self):
        return self.series_data.index
#データの切り出し
    def next_batch(self, input_digits, output_digits, ample):
        data = []
        target = []

        n_index = self.series_length - (input_digits + output_digits)
        noise = ample * np.random.uniform(low=-1.0, high=1.0, \
            size=[self.series_length, self.feature_count])
        value = self.series_data.values
        noise_value = value + noise

        for i in range(0, n_index):

            data.append(noise_value[i: i + input_digits])
            target.append(spy.zscore(value[i+input_digits: i+input_digits+output_digits]))

        X = np.stack(data)
        std_Y = np.stack(target)

        #import pdb; pdb.set_trace()
        return X, std_Y

    def append(self, data_point):
        dataframe = pd.DataFrame(data_point, columns=self.series_data.columns)
        self.series_data = self.series_data.append(dataframe)

    def tail(self, n):
        return TimeSeriesDataSet(self.series_data.tail(n))

    def as_array(self):
        return np.stack([self.series_data.as_matrix()])
    #標準化
    def mean(self):
        return self.series_data.mean()

    def std(self):
        return self.series_data.std()

    def standardize(self, mean = None, std = None):
        if mean is None:
            mean = self.mean()
        if std is None:
            std = self.std()
        return TimeSeriesDataSet((self.series_data - mean) / std)


class Early_Stopping():
    def __init__(self, patience=0, verbose=0):
        self._step = 0
        self._loss = float('inf')
        self.patience = patience
        self.verbose = verbose

    def validate(self, loss):
        if self._loss < loss:
            self._step += 1
            if self._step > self.patience:
                if self.verbose:
                    print('early stopping')
                return True
        else:
            self._step = 0
            self._loss = loss

        return False

eval_data_set_kari = pd.read_csv(input_data_path)
eval_data_set = eval_data_set_kari[['date_time', 'time_diff']]
eval_data_set = eval_data_set.set_index(['date_time'])
eval_data_set_inst = TimeSeriesDataSet(eval_data_set)
eval_series_length = eval_data_set_inst.series_length
(eval_X, eval_Y) = eval_data_set_inst.next_batch(input_digits = input_digits, \
    output_digits=output_digits, ample = ample)

n_in = len(eval_X[0][0])
n_out = len(eval_Y[0][0])

N_train = int((learning_data_day_len * output_digits - (input_digits + output_digits))* 0.95)
N_validation = (learning_data_day_len * output_digits- (input_digits + output_digits)) - N_train
n_batches = N_train // batch_size

num_day = eval_series_length // output_digits - 1

dataframe_2_  = []
day_d = []
series_error = []
gauss_error = []
log_gauss_error = []
rnn_day_series_mape = []
p_data_sr = []
rnn_np_p_data_sr = np.empty(0, int)
an_d = 1
check = 0
num_of_err = 0

anom_day_fig = plt.figure(figsize=(15, 25))

# for k in range(0, (eval_series_length - (learning_data_day_len * output_digits + output_digits)) // output_digits - 1):
k=1
tf.reset_default_graph()
def inference(x, y, n_batch, is_training,
              input_digits=None,
              output_digits=None,
              n_hidden=None,
              n_out=None):
    def weight_variable(shape):
        initial = tf.truncated_normal(shape, stddev=0.01)
        return tf.Variable(initial)

    def bias_variable(shape):
        initial = tf.zeros(shape, dtype=tf.float32)
        return tf.Variable(initial)

    def batch_normalization(shape, x):
        with tf.name_scope('batch_normalization'):
            eps = 1e-8
            # beta = tf.Variable(tf.zeros(shape))
            # gamma = tf.Variable(tf.ones(shape))
            mean, var = tf.nn.moments(x, [0, 1])
            # nom_batch = gamma * (x - mean) / tf.sqrt(var + eps) + beta
            nom_batch = (x - mean) / tf.sqrt(var + eps)
            # print(nom_batch[0], len(nom_batch[0]))
            return nom_batch

    encoder = cudnn_rnn.CudnnGRU(
        num_layers=1,
        num_units=int(n_hidden),
        input_mode='auto_select',
        direction='bidirectional',
        dtype=tf.float32)

        # state = encoder._zero_state(n_batches)

        # [input_digits, n_batch, 1], [1, n_batch, n_hidden]
    encoder_outputs, encoder_states = \
        encoder(tf.reshape(batch_normalization(input_digits, x), \
                [input_digits, n_batch, n_in]),
            # initial_state = state,
            training = True
        )

    encoder_states_fw = tf.slice(encoder_states, [0, 0, 0, 0], [1, 1, n_batch, n_hidden])
    encoder_states_fw = tf.reshape(encoder_states_fw, [n_batch, n_hidden])
    # encoder_forward = rnn_cell.GRUCell(n_hidden, reuse=tf.AUTO_REUSE)
    # encoder_backward = rnn_cell.GRUCell(n_hidden, reuse=tf.AUTO_REUSE)
    # encoder_outputs = []
    # encoder_states = []
    #
    # # size = [batch_size][input_digits][input_len]
    # x = tf.transpose(batch_normalization(input_digits, x), [1, 0, 2])
    # x = tf.reshape(x, [-1, n_in])
    # x = tf.split(x, input_digits, 0)
    # # Encode
    #
    # # state = encoder.zero_state(n_batch, tf.float32)
    #
    # # with tf.variable_scope('Encoder'):
    # #     for t in range(input_digits):
    # #         if t > 0:
    # #             tf.get_variable_scope().reuse_variables()
    # #         (output, state) = encoder(batch_normalization(input_digits, x)[:, t, :], state)
    # #         encoder_outputs.append(output)
    # #         encoder_states.append(state)
    #
    # encoder_outputs, encoder_states_fw, encoder_states_bw = tf.nn.static_bidirectional_rnn(
    #     encoder_forward,
    #     encoder_backward,
    #     x,
    #     dtype=tf.float32
    #     )
    # encoder_outputs size = [time][batch][cell_fw.output_size + cell_bw.output_size]
    # encoder_states_fw, encoder_states_bw is final state
    # Decode


    AttentionMechanism = seq2seq.BahdanauAttention(num_units=num_units,
                                                    memory=tf.reshape(encoder_outputs, \
                                                        [n_batch, input_digits, n_hidden * 2])
                                                    )
                                                    # when use bidirectional, n_hidden * 2
                                                    # tf.reshape(encoder_outputs, n_batch, input_digits, ),
                                                    # memory_sequence_length = input_digits)
                                                    # normalize=True)


    decoder_1 = rnn_cell.GRUCell(n_hidden, reuse = tf.AUTO_REUSE)
    # decoder_2 = rnn_cell.GRUCell(n_hidden, reuse = tf.AUTO_REUSE)

    decoder_1= seq2seq.AttentionWrapper(decoder_1,
                                       attention_mechanism = AttentionMechanism,
                                       attention_layer_size = attention_layer_size,
                                       output_attention = False)
                                       # initial_cell_state = encoder_states[-1])こいつが悪い

    # decoder_2= seq2seq.AttentionWrapper(decoder_2,
    #                                    attention_mechanism = AttentionMechanism,
    #                                    attention_layer_size = 50,
    #                                    output_attention = False,
    #                                    name = 'att_lay_2')

    state_1 = decoder_1.zero_state(n_batch, tf.float32)\
        .clone(cell_state=encoder_states_fw)

    # state_2 = decoder_2.zero_state(n_batch, tf.float32)
        # .clone(cell_state=tf.reshape(encoder_states_bw[-1], [n_batch, n_hidden]))

    # state = encoder_states[-1]
    # decoder_outputs = tf.reshape(encoder_outputs[-1,　:,　:], [n_batch, 1])
    # [input_len, n_batch, n_hidden]
    decoder_1_outputs = tf.slice(encoder_outputs, [input_digits-2, 0, 0], [1, n_batch, n_hidden])
    # decoder_2_outputs = tf.slice(encoder_outputs, [input_digits-2, 0, n_hidden], [1, n_batch, n_hidden])
    # decoder_2_outputs = encoder_outputs[:, :, n_hidden:][-1]
    # decoder_outputs = [encoder_outputs[-1]]

    # 出力層の重みとバイアスを事前に定義
    V_hid_1 = weight_variable([n_hidden, n_out])
    c_hid_1 = bias_variable([n_out])

    V_hid_2 = weight_variable([n_hidden, n_out])
    c_hid_2 = bias_variable([n_out])

    V_out = weight_variable([n_hidden, n_out])
    c_out = bias_variable([n_out])

    fc_outputs = []

    # decoder = seq2seq.BasicDecoder(cell = decoder,
    #                                 heiper = helper,
    #                                 initial_state=state,
    #                                 )

    elems = tf.convert_to_tensor([1, 0])
    samples = tf.multinomial(tf.log([[tchr_frcng_thr, 1 - tchr_frcng_thr]]), 1) # note log-prob

    with tf.variable_scope('Decoder'):
        for t in range(1, output_digits):
            if t > 1:
                tf.get_variable_scope().reuse_variables()
                # tf.get_variable_scope().reuse_variables()


            if is_training is True:
                cell_input_bin = elems[tf.cast(samples[0][0], tf.int32)]
                # bool = tf.equal(cell_input_bin, 1)
                t_const = tf.const(t)
                cell_input = tf.case({tf.equal(cell_input_bin, 1): lambda: batch_normalization(output_digits, y)[:, t-1, :],
                        tf.equal(t_const, 1): lambda: tf.matmul(decoder_1_outputs[-1], V_hid_1) + c_hid_1},
                    default=lambda: output_1)
                # cell_input_bin = np.randam.choice([1, 0],p=[tchr_frcng_thr, 1 - tchr_frcng_thr])
                #
                # if cell_input_bin==1:
                #     cell_input = batch_normalization(output_digits, y)[:, t-1, :]
                #
                # elif t == 1:
                #     cell_input = tf.matmul(decoder_1_outputs[-1], V_hid_1) + c_hid_1
                #
                # else:
                #     cell_input = output_1

                (output_1, state_1) = decoder_1(cell_input, state_1)
                # (output_2, state_2) = decoder_2(batch_normalization(output_digits, y)[:, t-1, :], state_2)
            else:
                # 直前の出力を求める
                out_1 = tf.matmul(decoder_1_outputs[-1], V_hid_1) + c_hid_1#to hidden layer
                # out_2 = tf.matmul(decoder_2_outputs[-1], V_hid_2) + c_hid_2#to hidden layer
                # fc_out = tf.matmul(tf.concat([decoder_1_outputs[-1], decoder_2_outputs[-1]], 1), V_out) + c_out
                #forecast data

                # elems = decoder_outputs[-1], V , c
                # out = tf.map_fn(lambda x: x[0] * x[1] + x[2], elems)
                # out = decoder_outputs
                fc_outputs.append(out_1)
                (output_1, state_1) = decoder_1(out_1, state_1)
                # (output_2, state_2) = decoder_2(out_2, state_2)

            # decoder_outputs.append(output)
            decoder_1_outputs = tf.concat([decoder_1_outputs, tf.reshape(output_1, [1, n_batch, n_hidden])], axis = 0)
            # decoder_2_outputs = tf.concat([decoder_2_outputs, tf.reshape(output_2, [1, n_batch, n_hidden])], axis = 0)
            # decoder_outputs = tf.concat([decoder_outputs, output], 1)
    if is_training is True:
        output = tf.reshape(tf.concat(decoder_1_outputs, axis=1),
                            [-1, output_digits, n_hidden])
        with tf.name_scope('check'):
            linear = tf.einsum('ijk,kl->ijl', output, V_out, ) + c_out
            return linear
    else:
        # 最後の出力を求める
        fc_out = tf.matmul(tf.concat(decoder_1_outputs[-1], 1), V_out) + c_out
        fc_outputs.append(fc_out)

        output = tf.reshape(tf.concat(fc_outputs, axis=1),
                            [-1, output_digits, n_out])
        return output

def loss(y, t):
    with tf.name_scope('loss'):
        # mse = tf.reduce_mean(tf.square(y - t), axis = [1, 0])
        norm = tf.contrib.distributions.Normal(0., 0.5)
        error = y-t
        pdf = norm.prob(error)
        loss= tf.reduce_mean(1 - pdf, [1, 0])
        # mse = tf.reduce_mean(tf.square(y - t), [1, 0])
        return loss

def training(loss, learning_rate):
    with tf.name_scope('train_step'):
        optimizer = \
            tf.train.AdamOptimizer(learning_rate=learning_rate, beta1=0.9, beta2=0.999)

        train_step = optimizer.minimize(loss)
        return train_step


x = tf.placeholder(tf.float32, shape=[None, input_digits, n_in])
t = tf.placeholder(tf.float32, shape=[None, output_digits, n_out])
n_batch = tf.placeholder(tf.int32, shape=[])
is_training = tf.placeholder(tf.bool)

y = inference(x, t, n_batch, is_training,
              input_digits=input_digits,
              output_digits=output_digits,
              n_hidden=n_hidden, n_out=n_out)
loss = loss(y, t)
train_step = training(loss = loss, learning_rate = learning_rate)

with tf.name_scope('initial'):
    init = tf.global_variables_initializer()
    sess = tf.Session()
    # sess = tf_debug.LocalCLIDebugWrapperSession(sess)
    # sess = tf_debug.TensorBoardDebugWrapperSession(sess, 'localhost:6064')
    if k == 0:
        tf.summary.FileWriter(LOG_DIR, sess.graph)
    sess.run(init)

history = {
    'val_loss': [],
    'val_acc': []
}

input_data = eval_X[k * output_digits: (k + learning_data_day_len - 1) * output_digits]
true_data = eval_Y[k * output_digits: (k + learning_data_day_len - 1) * output_digits]

# print(input_data[:3])
input_data_train, input_data_validation, true_data_train, \
    true_data_validation = train_test_split(input_data, true_data, \
        test_size = N_validation)
# size = [batch_size][input_digits][input_len]
early_stopping = Early_Stopping(patience=10, verbose=1)

# print('input_data_train = ', input_data_train[1:4])
# print(len(input_data_train[0]))
for epoch in range(epochs):
    X_, Y_ = shuffle(input_data_train, true_data_train)

    with tf.name_scope('train'):
        for h in range(100):
            start = h * batch_size
            end = start + batch_size
            # print('begin learning')
            # print(y)
            # t = Y_[start:end]
            # y = inference(X_[start:end], Y_[start:end], batch_size, True,
            #               input_digits=input_digits,
            #               output_digits=output_digits,
            #               n_hidden=n_hidden, n_out=n_out)
            # loss = loss(y, t)
            #
            # train_step(loss, learning_rate)
            sess.run(train_step, feed_dict={
                x: X_[start:end],
                t: Y_[start:end],
                n_batch: batch_size,
                is_training: True
            })

    val_loss = loss.eval(session=sess, feed_dict={
        x: input_data_validation,
        t: true_data_validation,
        n_batch: N_validation,
        is_training: False
    })

    #mean_val, var_val = tf.nn.moments(X_validation, [0, 1])
    #std_val_loss = val_loss * tf.sqrt(var_val) + mean_val
    #std_val_loss = val_loss / train_std**2

    history['val_loss'].append(val_loss)
    print('epoch:', epoch,
          ' validation loss:', val_loss)

    if val_loss < 0.05 and epoch == eary_stop_epoch:
        break

#forcasting
predicted_traffic = [[None] * len(eval_data_set.columns) \
for l in range(input_digits)]

fc_input = eval_X[learning_data_day_len * output_digits - (input_digits - k * output_digits)].reshape(1, input_digits, 1)
std_fc_input = spy.zscore(fc_input, axis = 1)

z_ = std_fc_input.reshape(1, input_digits, 1)

std_output = y.eval(session=sess, feed_dict={
    x: z_,
    n_batch: 1,
    is_training: False
})

tf.reset_default_graph()

fc_input_mean = fc_input.mean(axis=1, keepdims=True)
fc_input_std = fc_input.std(axis=1, keepdims=True)

fc_output = std_output * fc_input_std + fc_input_mean
fc_seq = fc_output.reshape(-2)
print('forecasting', fc_seq)
rnn_np_p_data_sr = np.append(rnn_np_p_data_sr, fc_seq.reshape(-1), axis = 0)

dataframe_2_ = eval_data_set[(learning_data_day_len + k) * output_digits: \
    (learning_data_day_len + k) * output_digits + output_digits]
day_d = dataframe_2_.values.reshape(-1)
# print(day_d)

# if len(day_d) != output_digits:
#     break

# series_error.append(fc_seq - day_d)
# print(series_error)
# print(gauss_error)
# log_gauss_error.append(np.log10(gauss_error[k]))

predicted_traffic_data = pd.DataFrame(rnn_np_p_data_sr)
    # columns = eval_data_set[learning_data_day_len * output_digits:learning_data_day_len * output_digits + len(rnn_np_p_data_sr)].columns, \
    # index=eval_data_set[learning_data_day_len * output_digits:learning_data_day_len * output_digits + len(rnn_np_p_data_sr)].index)

predicted_traffic_data.to_excel(path_output_data + 'interval.xlsx')

# wb = openpyxl.load_workbook(path_output_data + '/seq2seq_predict.xlsx')
# sheet = wb['Sheet1']
# sheet.cell(row=1, column=5, value='real_number')
# real_day_data = eval_data_set.values.reshape(-1)[learning_data_day_len * output_digits:]
# for i in range(len(real_day_data)):
#     sheet.cell(row = 2 + i, column=5, value=real_day_data[i])
# wb.save(path_output_data + '/seq2seq_predict.xlsx')
